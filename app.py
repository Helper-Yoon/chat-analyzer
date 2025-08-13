import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import io

# 페이지 설정
st.set_page_config(
    page_title="SNS센터 채팅분석",
    page_icon="📊",
    layout="wide"
)

# 컴팩트한 스타일
st.markdown("""
    <style>
    /* 상단 패딩 제거 */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 1rem;
        max-width: 1000px;
    }
    
    /* 제목 바 스타일 */
    .title-bar {
        background: linear-gradient(90deg, #004C99 0%, #0066CC 100%);
        color: white;
        padding: 0.8rem 1.5rem;
        border-radius: 5px;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }
    
    .title-bar h1 {
        margin: 0;
        font-size: 1.5rem;
        font-weight: bold;
    }
    
    /* 분석 실행 버튼 스타일 */
    .stButton > button {
        background: linear-gradient(90deg, #004C99 0%, #0066CC 100%);
        color: white;
        width: 100%;
        height: 50px;
        font-size: 18px;
        font-weight: bold;
        border: none;
        border-radius: 5px;
        margin-top: 1rem;
    }
    
    .stButton > button:hover {
        background: linear-gradient(90deg, #0066CC 0%, #0080FF 100%);
    }
    
    /* 완료 페이지 스타일 */
    .success-container {
        text-align: center;
        padding: 2rem;
        background: #f0f8ff;
        border-radius: 10px;
        border: 2px solid #004C99;
        margin: 2rem 0;
    }
    
    .success-icon {
        font-size: 4rem;
        color: #004C99;
        margin-bottom: 1rem;
    }
    
    /* 입력 필드 스타일 */
    .stTextInput > div > div > input {
        border: 1px solid #d0d0d0;
        border-radius: 4px;
        padding: 0.4rem;
    }
    
    /* 섹션 제목 스타일 */
    h5 {
        color: #004C99;
        border-bottom: 2px solid #e0e0e0;
        padding-bottom: 0.5rem;
        margin-bottom: 0.8rem;
        margin-top: 1rem;
    }
    
    /* 컬럼 내 라벨 스타일 */
    .label-text {
        font-weight: 600;
        color: #333;
        display: flex;
        align-items: center;
        height: 38px;
    }
    
    /* 섹션 간격 줄이기 */
    .element-container {
        margin-bottom: 0.5rem !important;
    }
    </style>
""", unsafe_allow_html=True)

class CollaborationAnalyzer:
    def __init__(self):
        if 'analysis_complete' not in st.session_state:
            st.session_state.analysis_complete = False
        if 'result_file' not in st.session_state:
            st.session_state.result_file = None
        if 'show_result_page' not in st.session_state:
            st.session_state.show_result_page = False

    def load_and_process_data(self, file, start_date_str, end_date_str):
        try:
            # 엑셀 파일 로딩
            all_sheets = pd.read_excel(file, sheet_name=None, engine='openpyxl')
            
            required_sheets = ['UserChat data', 'Message data', 'Manager data']
            sheet_data = {core_name: [] for core_name in required_sheets}
            
            for sheet_name, df in all_sheets.items():
                for core_name in required_sheets:
                    if core_name in sheet_name:
                        sheet_data[core_name].append(df)
            
            if not all(sheet_data.values()):
                st.error("필수 시트를 찾을 수 없습니다.")
                return None

            # 시트 통합
            user_chat_df = pd.concat(sheet_data['UserChat data'], ignore_index=True).drop_duplicates(subset=['id'])
            message_df = pd.concat(sheet_data['Message data'], ignore_index=True).drop_duplicates(subset=['chatId', 'personId', 'createdAt', 'plainText'])
            manager_df = pd.concat(sheet_data['Manager data'], ignore_index=True).drop_duplicates(subset=['id'])

            # ID 정제
            def clean_id(series):
                return series.astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

            user_chat_df['id'] = clean_id(user_chat_df['id'])
            user_chat_df['assigneeId'] = clean_id(user_chat_df['assigneeId'])
            message_df['chatId'] = clean_id(message_df['chatId'])
            message_df['personId'] = clean_id(message_df['personId'])
            manager_df['id'] = clean_id(manager_df['id'])
            
            # 날짜 필터링
            start_ts = pd.to_datetime(start_date_str)
            end_ts = pd.to_datetime(end_date_str) + pd.DateOffset(days=1)

            user_chat_df['firstOpenedAt'] = pd.to_datetime(user_chat_df['firstOpenedAt'], errors='coerce')
            user_chat_df.dropna(subset=['firstOpenedAt'], inplace=True)
            filtered_user_chat_df = user_chat_df[(user_chat_df['firstOpenedAt'] >= start_ts) & (user_chat_df['firstOpenedAt'] < end_ts)]

            message_df['createdAt'] = pd.to_datetime(message_df['createdAt'], errors='coerce')
            message_df.dropna(subset=['createdAt'], inplace=True)
            filtered_message_df = message_df[(message_df['createdAt'] >= start_ts) & (message_df['createdAt'] < end_ts)]
            
            if filtered_message_df.empty:
                st.error("선택된 기간 내에 데이터가 없습니다.")
                return None
            
            # 데이터 병합
            merged_df = pd.merge(filtered_message_df, user_chat_df[['id', 'assigneeId']], left_on='chatId', right_on='id', how='left').dropna(subset=['assigneeId'])
            merged_df = pd.merge(merged_df, manager_df[['id', 'name']], left_on='personId', right_on='id', how='left', suffixes=('', '_manager')).rename(columns={'name': 'authorName'}).dropna(subset=['authorName'])
            
            return {'merged': merged_df, 'user_chat': filtered_user_chat_df, 'manager': manager_df}

        except Exception as e:
            st.error(f"데이터 처리 중 오류: {str(e)}")
            return None

    def create_output_excel(self, processed_data, start_date_str, end_date_str, managers_list, exclusion_list):
        df_merged = processed_data['merged']
        user_chat_df = processed_data['user_chat']
        manager_df = processed_data['manager']
        
        output = io.BytesIO()
        
        # 데이터 분류
        manager_data = df_merged[df_merged['authorName'].isin(managers_list)]
        agent_data = df_merged[(~df_merged['authorName'].isin(managers_list)) & (~df_merged['authorName'].isin(exclusion_list))]

        all_agents = pd.DataFrame({'authorName': agent_data['authorName'].unique()})
        agent_non_assignee = agent_data[agent_data['personId'] != agent_data['assigneeId']]
        
        if not agent_non_assignee.empty:
            collaborated_chats = agent_non_assignee.groupby('authorName')['chatId'].nunique()
        else:
            collaborated_chats = pd.Series(dtype='int64', name='chatId', index=pd.Index([], name='authorName'))

        total_chats = agent_data.groupby('authorName')['chatId'].nunique()
        hir_summary = (collaborated_chats / total_chats).reset_index(name='HIR').fillna(0)
        
        total_msg_counts = agent_data.groupby('authorName').size().reset_index(name='total_messages')
        
        # 필터링
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        analysis_days = (end_date - start_date).days + 1
        min_msg_threshold = analysis_days * 10

        filter_df = pd.merge(all_agents, hir_summary, on='authorName', how='left')
        filter_df = pd.merge(filter_df, total_msg_counts, on='authorName', how='left')
        filter_df.fillna(0, inplace=True)

        filtered_authors_df = filter_df[
            (filter_df['HIR'] > 0) & (filter_df['HIR'] < 1) & 
            (filter_df['total_messages'] > 10) &
            (filter_df['total_messages'] >= min_msg_threshold)
        ]
        
        if filtered_authors_df.empty:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame({'알림': ['필터 조건에 해당하는 상담사가 없어 데이터를 생성할 수 없습니다.']}).to_excel(writer, sheet_name='결과 없음', index=False)
            output.seek(0)
            return output

        authors_to_keep = filtered_authors_df['authorName'].unique()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            def analyze_group(group_name, group_merged_df, user_chat_df_local, manager_df_local):
                if group_merged_df.empty and group_name == "상담사":
                    return pd.DataFrame(), pd.DataFrame()

                authors = pd.DataFrame({'authorName': pd.concat([group_merged_df['authorName'], manager_df_local.loc[manager_df_local['id'].isin(user_chat_df_local['assigneeId']), 'name']]).unique()})
                
                group_non_assignee_df = group_merged_df[group_merged_df['personId'] != group_merged_df['assigneeId']].copy()
                group_assignee_df = group_merged_df[group_merged_df['personId'] == group_merged_df['assigneeId']].copy()

                metrics_df = authors.copy()
                if not group_non_assignee_df.empty:
                    # HIR 계산
                    hir_s = (group_non_assignee_df.groupby('authorName')['chatId'].nunique() / group_merged_df.groupby('authorName')['chatId'].nunique()).reset_index(name='HIR').fillna(0)
                    
                    # IIF 계산
                    chat_lengths = group_merged_df.groupby('chatId').size().to_dict()
                    group_non_assignee_df['chat_length'] = group_non_assignee_df['chatId'].map(chat_lengths)
                    iif_s = group_non_assignee_df.groupby('authorName')['chat_length'].sum().reset_index(name='IIF')
                    
                    # CIS 계산
                    core_keywords = ['월요금', '사은품', '위약금', '결합', '설치일', '설치비', '약정', '지원금', '할인', '통신사', '요금제', '인터넷', '휴대폰']
                    keyword_pattern = '|'.join(core_keywords)
                    group_non_assignee_df['cis_flag'] = group_non_assignee_df['plainText'].str.contains(keyword_pattern, na=False).astype(int)
                    cis_s = group_non_assignee_df.groupby('authorName')['cis_flag'].sum().reset_index(name='CIS')

                    # DLS 계산
                    group_non_assignee_df['msg_length'] = group_non_assignee_df['plainText'].str.len()
                    dls_s = group_non_assignee_df.groupby('authorName')['msg_length'].mean().reset_index(name='DLS')
                    
                    # ALS 계산
                    group_non_assignee_df['als_flag'] = group_non_assignee_df['plainText'].str.contains('https://form.ajd.co.kr/', na=False).astype(int)
                    als_s = group_non_assignee_df.groupby('authorName')['als_flag'].sum().reset_index(name='ALS')
                    als_s['ALS'] *= 10

                    for df in [hir_s, iif_s, cis_s, dls_s, als_s]:
                        metrics_df = pd.merge(metrics_df, df, on='authorName', how='left')
                metrics_df.fillna(0, inplace=True)

                # 통계 계산
                assigned_stats = group_assignee_df.groupby('authorName').agg(
                    담당_메시지_수=('chatId', 'size'),
                    담당_글자_수=('plainText', lambda x: x.str.len().sum())
                ).reset_index()
                
                help_stats = group_non_assignee_df.groupby('authorName').agg(
                    도움_상담_수=('chatId', 'nunique'),
                    도움_메시지_수=('chatId', 'size'),
                    도움_글자_수=('plainText', lambda x: x.str.len().sum())
                ).reset_index()
                
                author_ids = manager_df_local[manager_df_local['name'].isin(authors['authorName'])]['id']
                assigned_chat_counts = user_chat_df_local[user_chat_df_local['assigneeId'].isin(author_ids)].groupby('assigneeId').agg(
                    담당_상담_수=('id', 'nunique')
                ).reset_index()
                assigned_chat_counts = pd.merge(assigned_chat_counts, manager_df_local[['id', 'name']], left_on='assigneeId', right_on='id').rename(columns={'name': 'authorName'})

                # 정성 점수 계산
                def get_base_score(df):
                    if df.empty:
                        return pd.DataFrame(columns=['authorName', 'base_score'])
                    message_counts = df['plainText'].value_counts()
                    df['score'] = df['plainText'].map(lambda x: np.log1p(len(df) / message_counts.get(x, 1)))
                    return df.groupby('authorName')['score'].sum().reset_index(name='base_score')

                base_help_score = get_base_score(group_non_assignee_df)
                base_assigned_score = get_base_score(group_assignee_df)

                # 최종 데이터 병합
                summary_df = pd.merge(authors, assigned_chat_counts[['authorName', '담당_상담_수']], on='authorName', how='left')
                summary_df = pd.merge(summary_df, assigned_stats, on='authorName', how='left')
                summary_df = pd.merge(summary_df, help_stats, on='authorName', how='left')
                summary_df = pd.merge(summary_df, base_help_score.rename(columns={'base_score': 'base_help_score'}), on='authorName', how='left')
                summary_df = pd.merge(summary_df, base_assigned_score.rename(columns={'base_score': 'base_assigned_score'}), on='authorName', how='left')
                summary_df = pd.merge(summary_df, metrics_df, on='authorName', how='left')
                summary_df.fillna(0, inplace=True)

                # 보정 계산
                metrics_for_correction = ['ALS']
                for col in metrics_for_correction:
                    min_val, max_val = summary_df[col].min(), summary_df[col].max()
                    if max_val > min_val:
                        summary_df[f'norm_{col}'] = (summary_df[col] - min_val) / (max_val - min_val)
                    else:
                        summary_df[f'norm_{col}'] = 0
                
                als_weight = 3
                summary_df['help_correction'] = summary_df.get('norm_ALS', 0) * als_weight
                summary_df['assigned_correction'] = 0
                
                summary_df['도움_정성_점수'] = summary_df.get('base_help_score', 0) * (1 + summary_df['help_correction'])
                summary_df['담당_정성_점수'] = summary_df.get('base_assigned_score', 0) * (1 + summary_df['assigned_correction'])

                total_help_score = summary_df['도움_정성_점수'].sum()
                total_assigned_score = summary_df['담당_정성_점수'].sum()
                if total_assigned_score > 0 and total_help_score > 0:
                    ratio_factor = (total_help_score / 3) / total_assigned_score
                    summary_df['담당_정성_점수'] *= ratio_factor
                
                summary_df['총_정성_점수'] = summary_df['도움_정성_점수'] + summary_df['담당_정성_점수']

                # 최종 컬럼 정리
                final_cols = ['authorName', '담당_상담_수', '담당_메시지_수', '담당_글자_수', '도움_상담_수', '도움_메시지_수', '도움_글자_수', '담당_정성_점수', '도움_정성_점수', '총_정성_점수']
                final_summary = summary_df[final_cols].rename(columns={
                    'authorName': '상담사명', '담당_상담_수': '담당 상담 수', '담당_메시지_수': '담당 메시지 수',
                    '담당_글자_수': '담당 글자 수', '도움_상담_수': '도움 상담 수', '도움_메시지_수': '도움 메시지 수',
                    '도움_글자_수': '도움 글자 수', '담당_정성_점수': '담당 정성 점수', '도움_정성_점수': '도움 정성 점수',
                    '총_정성_점수': '총 정성 점수'
                })
                
                metrics_cols = ['authorName', 'HIR', 'IIF', 'CIS', 'DLS', 'ALS']
                final_metrics = summary_df[metrics_cols].rename(columns={
                    'authorName': '상담사명', 'HIR': '도움 개입률', 'IIF': '개입 영향력 계수',
                    'CIS': '콘텐츠 정보 점수', 'DLS': '언어 깊이 점수', 'ALS': '신청서 링크 점수'
                })

                return final_summary.round(2), final_metrics.round(2)

            # 그룹별 분석
            agent_summary_df, agent_metrics_df = analyze_group(
                "상담사",
                agent_data[agent_data['authorName'].isin(authors_to_keep)],
                user_chat_df[user_chat_df['assigneeId'].isin(manager_df[manager_df['name'].isin(authors_to_keep)]['id'])],
                manager_df
            )
            
            manager_summary_df, _ = analyze_group(
                "관리자",
                manager_data,
                user_chat_df[user_chat_df['assigneeId'].isin(manager_df[manager_df['name'].isin(managers_list)]['id'])],
                manager_df
            )
            
            # 엑셀 시트 생성
            if not agent_summary_df.empty:
                agent_summary_df = agent_summary_df.sort_values(by='총 정성 점수', ascending=False)
                agent_summary_df.to_excel(writer, sheet_name='채팅분석_요약', index=False)
            
            if not manager_summary_df.empty:
                manager_summary_df = manager_summary_df.sort_values(by='총 정성 점수', ascending=False)
                manager_summary_df.to_excel(writer, sheet_name='관리자_분석', index=False)
            
            if not agent_metrics_df.empty:
                agent_metrics_df = agent_metrics_df.sort_values(by='신청서 링크 점수', ascending=False)
                agent_metrics_df.to_excel(writer, sheet_name='채팅분석_지표', index=False)

            # 스코어보드 생성
            workbook = writer.book
            worksheet = workbook.create_sheet('스코어보드')
            
            metrics_to_rank = {
                '담당 상담 수': '담당 상담 수', '담당 메시지 수': '담당 메시지 수', 
                '도움 상담 수': '도움 상담 수', '도움 메시지 수': '도움 메시지 수',
                '총 정성 점수': '총 정성 점수'
            }
            
            current_row = 1
            for col, title in metrics_to_rank.items():
                if not agent_summary_df.empty:
                    top_5 = agent_summary_df.nlargest(5, col)[['상담사명', col]]
                    
                    worksheet.cell(row=current_row, column=1, value=f'--- {title} Top 5 ---')
                    current_row += 1
                    
                    for i in range(len(top_5)):
                        worksheet.cell(row=current_row + i, column=1, value=f"{i+1}위")
                        worksheet.cell(row=current_row + i, column=2, value=top_5.iloc[i, 0])
                        worksheet.cell(row=current_row + i, column=3, value=top_5.iloc[i, 1])
                    
                    current_row += 6

        output.seek(0)
        return output

# 메인 앱
def main():
    analyzer = CollaborationAnalyzer()
    
    # 제목 바
    st.markdown("""
        <div class="title-bar">
            <div style="display: flex; align-items: center;">
                <span style="font-size: 1.8rem; margin-right: 10px;">📊</span>
                <h1>SNS센터 채팅분석 프로그램 v1.9</h1>
            </div>
            <span style="font-size: 0.9rem;">채팅 데이터 협업 성과 분석</span>
        </div>
    """, unsafe_allow_html=True)
    
    # 분석 완료 페이지
    if st.session_state.show_result_page:
        st.markdown("""
            <div class="success-container">
                <div class="success-icon">✅</div>
                <h2 style="color: #004C99; margin-bottom: 1rem;">분석이 완료되었습니다!</h2>
                <p style="color: #666; margin-bottom: 2rem;">
                    분석 결과가 준비되었습니다.<br>
                    아래 버튼을 클릭하여 Excel 파일을 다운로드하세요.
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # 다운로드 섹션
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SNS센터_채팅분석_결과_{timestamp}.xlsx"
            
            st.download_button(
                label="📥 결과 파일 다운로드",
                data=st.session_state.result_file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
            if st.button("🔄 새로운 분석 시작", use_container_width=True):
                st.session_state.show_result_page = False
                st.session_state.analysis_complete = False
                st.session_state.result_file = None
                st.rerun()
        
        # 생성된 시트 정보
        st.info("""
            📄 **생성된 시트**: 
            스코어보드 | 채팅분석_요약 | 관리자_분석 | 채팅분석_지표
        """)
    
    # 메인 분석 페이지
    else:
        # 1. 파일 업로드 (한 줄)
        st.markdown("##### 📁 파일 업로드")
        uploaded_file = st.file_uploader(
            "Excel 파일을 선택하세요",
            type=['xlsx'],
            help="UserChat, Message, Manager data 시트가 포함된 파일",
            label_visibility="collapsed"
        )
        if uploaded_file:
            st.success(f"✅ 업로드 완료: {uploaded_file.name}")
        
        # 2. 분석 기간 (한 줄)
        st.markdown("##### 📅 분석 기간")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            start_date = st.date_input(
                "시작일",
                value=datetime(2025, 7, 1)
            )
        with col2:
            end_date = st.date_input(
                "종료일", 
                value=datetime.now() - timedelta(days=1)
            )
        with col3:
            days = (end_date - start_date).days + 1
            st.metric("분석 일수", f"{days}일")
        
        # 3. 인원 설정 (두 줄)
        st.markdown("##### 👥 인원 설정")
        
        # 관리자 (첫 번째 줄)
        col1, col2 = st.columns([1, 9])
        with col1:
            st.markdown("<div class='label-text'>관리자</div>", unsafe_allow_html=True)
        with col2:
            managers = st.text_input(
                "관리자 목록",
                value="이민주, 이종민, 윤도우리, 김시진, 손진우",
                placeholder="관리자 이름을 쉼표로 구분하여 입력",
                label_visibility="collapsed"
            )
        
        # 제외 인원 (두 번째 줄)
        col1, col2 = st.columns([1, 9])
        with col1:
            st.markdown("<div class='label-text'>제외</div>", unsafe_allow_html=True)
        with col2:
            exclusions = st.text_input(
                "제외 목록",
                value="채주은, 정용욱, 한승윤, 김종현",
                placeholder="제외할 이름을 쉼표로 구분하여 입력 (선택사항)",
                label_visibility="collapsed"
            )
        
        # 구분선
        st.markdown("---")
        
        # 분석 실행 버튼 (전체 너비 파란 바)
        analyze_button = st.button(
            "🚀 분석 실행",
            type="primary",
            use_container_width=True,
            disabled=not uploaded_file
        )
        
        if not uploaded_file:
            st.warning("⚠️ 분석을 시작하려면 Excel 파일을 업로드해주세요.")
        
        # 분석 실행
        if analyze_button:
            with st.spinner("분석 중... 잠시만 기다려주세요"):
                # 프로그레스 바
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # 설정 파싱
                status_text.text("설정 확인 중...")
                progress_bar.progress(20)
                managers_list = [name.strip() for name in managers.split(',') if name.strip()]
                exclusion_list = [name.strip() for name in exclusions.split(',') if name.strip()]
                
                # 데이터 처리
                status_text.text("데이터 로딩 중...")
                progress_bar.progress(40)
                processed_data = analyzer.load_and_process_data(
                    uploaded_file,
                    start_date.strftime("%Y-%m-%d"),
                    end_date.strftime("%Y-%m-%d")
                )
                
                if processed_data:
                    # 결과 생성
                    status_text.text("분석 수행 중...")
                    progress_bar.progress(70)
                    
                    result_file = analyzer.create_output_excel(
                        processed_data,
                        start_date.strftime("%Y-%m-%d"),
                        end_date.strftime("%Y-%m-%d"),
                        managers_list,
                        exclusion_list
                    )
                    
                    status_text.text("결과 생성 중...")
                    progress_bar.progress(90)
                    
                    st.session_state.analysis_complete = True
                    st.session_state.result_file = result_file
                    st.session_state.show_result_page = True
                    
                    progress_bar.progress(100)
                    status_text.text("완료!")
                    
                    # 완료 페이지로 전환
                    st.rerun()

if __name__ == "__main__":
    main()
